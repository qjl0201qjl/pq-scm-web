from __future__ import annotations

import csv
import io
import json
import os
import re
import sqlite3
import time
from pathlib import Path
from typing import Any, Literal

import httpx
from fastapi import FastAPI, File, Form, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from pydantic import BaseModel

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "storage"
REPORT_DIR = DATA_DIR / "reports"
DB_PATH = DATA_DIR / "pq_scm.db"
DATA_DIR.mkdir(exist_ok=True)
REPORT_DIR.mkdir(exist_ok=True)

ASPECTS = ["外观造型", "内饰质感", "空间表现", "动力性能", "续航与能耗", "充电体验", "智能座舱", "智能驾驶", "操控与底盘", "舒适性与NVH", "安全配置", "售后服务与交付", "其他"]
TEXT_KEYS = ["text", "评论", "评论内容", "content", "review", "原始评论", "评论文本", "comment_text", "comment"]
IGNORED_TEXT_KEYS = {"index", "id", "comment_id", "label", "车型", "车系", "来源", "平台", "model", "source", "platform", "score", "user_score", "comment_score", "pub_time", "date", "price", "price_range"}


app = FastAPI(title="PQ-SCM API", version="1.0.0")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


class AbsaRunRequest(BaseModel):
    mode: Literal["rule", "llm", "hybrid"] = "hybrid"
    provider: str = "deepseek"
    batch_size: int = 20
    text_column: str | None = None


class DiagnosisRequest(BaseModel):
    top_n: int = 10


class QfdRequest(BaseModel):
    top_n: int = 10
    allow_llm_enhance: bool = False


class SupplyChainRequest(BaseModel):
    allow_llm_enhance: bool = False


class ReportRequest(BaseModel):
    vehicle_model: str | None = None
    start_date: str | None = None
    end_date: str | None = None
    report_type: str = "完整案例实证报告"
    top_n: int = 10
    output_format: Literal["excel", "word", "pdf"] = "excel"


def db() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db() -> None:
    with db() as conn:
        conn.executescript(
            """
            create table if not exists comments (
              comment_id text primary key,
              raw_text text not null,
              source text,
              vehicle_model text,
              time text
            );
            create table if not exists absa_results (
              comment_id text primary key,
              aspect text,
              opinion text,
              sentiment text,
              reason text,
              confidence real,
              need_review integer,
              analysis_source text
            );
            create table if not exists issue_summary (
              issue_id text primary key,
              issue_name text,
              aspect text,
              total_count integer,
              negative_count integer,
              positive_count integer,
              attention_A real,
              dissatisfaction_D real,
              intensity_I real,
              kano_type text,
              fda_score real,
              kano_factor real,
              final_PI real,
              evidence_json text
            );
            create table if not exists qfd_results (
              id integer primary key autoincrement,
              issue_id text,
              engineering_feature text,
              base_relation integer,
              keyword_match real,
              confidence_factor real,
              pi_factor real,
              relation_score real,
              module text,
              explanation text
            );
            create table if not exists supply_chain_results (
              id integer primary key autoincrement,
              engineering_feature text,
              module text,
              enterprise_name text,
              role_type text,
              collaboration_score real,
              collaboration_method text,
              recommendation_reason text,
              model_suggested integer default 0
            );
            create table if not exists report_records (
              report_id text primary key,
              report_type text,
              created_at text,
              file_path text
            );
            """
        )


init_db()


def rows(query: str, params: tuple[Any, ...] = ()) -> list[dict[str, Any]]:
    with db() as conn:
      return [dict(row) for row in conn.execute(query, params).fetchall()]


def norm_key(key: str) -> str:
    return re.sub(r"[\s_-]+", "", key.lower())


def looks_like_id(value: str) -> bool:
    return bool(re.match(r"^(NEV|EV|ID|NO|ROW)?[_-]?\d{3,}$", value.strip(), re.I))


def text_column_score(key: str, values: list[str]) -> float:
    nk = norm_key(key)
    if any(nk == norm_key(item) or nk in norm_key(item) for item in IGNORED_TEXT_KEYS):
        return -10000
    sample_values = [str(v).strip() for v in values if str(v).strip()][:80]
    sample = "".join(sample_values)
    chinese = len(re.findall(r"[\u4e00-\u9fa5]", sample))
    keywords = len(re.findall(r"续航|油耗|电耗|操控|动力|车机|座舱|中控|空间|内饰|外观|悬架|悬挂|减震|底盘|充电|智驾|雷达|售后|配置|舒适|噪声|异响|性价比|刹车|制动|座椅|空调|发动机", sample))
    avg_len = sum(len(v) for v in sample_values) / len(sample_values) if sample_values else 0
    long_ratio = sum(1 for v in sample_values if len(v) >= 8 and not looks_like_id(v)) / len(sample_values) if sample_values else 0
    id_ratio = sum(1 for v in sample_values if looks_like_id(v)) / len(sample_values) if sample_values else 1
    exact = 700 if any(nk == norm_key(item) for item in TEXT_KEYS) else 0
    partial = 260 if any(norm_key(item) in nk for item in TEXT_KEYS) else 0
    return exact + partial + chinese * 2 + keywords * 40 + avg_len * 4 + long_ratio * 220 - id_ratio * 800


def choose_text_column(data: list[dict[str, Any]]) -> tuple[str | None, bool]:
    if not data:
        return None, False
    keys = sorted({key for row in data for key in row.keys()})
    scored = [(key, text_column_score(key, [str(row.get(key, "")) for row in data])) for key in keys]
    scored.sort(key=lambda item: item[1], reverse=True)
    if scored and scored[0][1] > 0:
        return scored[0][0], scored[0][1] >= 260
    return None, False


def parse_upload(file_name: str, content: bytes) -> list[dict[str, Any]]:
    lower = file_name.lower()
    if lower.endswith(".xlsx") or lower.endswith(".xls"):
        try:
            from openpyxl import load_workbook
        except Exception as exc:
            raise HTTPException(500, f"Excel解析需要安装 openpyxl: {exc}")
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        values = list(sheet.values)
        if not values:
            return []
        header = [str(v or f"col_{i}") for i, v in enumerate(values[0])]
        return [{header[i]: value for i, value in enumerate(row)} for row in values[1:]]
    text = content.decode("utf-8-sig", errors="ignore")
    delimiter = "\t" if "\t" in text.splitlines()[0] else ","
    return list(csv.DictReader(io.StringIO(text), delimiter=delimiter))


def pick(row: dict[str, Any], keys: list[str], default: str = "") -> str:
    for key in row:
        if any(norm_key(candidate) in norm_key(key) for candidate in keys):
            return str(row.get(key) or default)
    return default


def classify_sentiment(text: str) -> str:
    if re.search(r"满意|喜欢|舒服|好看|省电|流畅|稳定|宽敞|不错|好用", text):
        return "positive"
    if re.search(r"不好|不满|卡顿|异响|掉电|缩水|太硬|投诉|误报|中断|严重|差|慢|崩溃", text):
        return "negative"
    return "neutral"


def rule_absa(text: str) -> dict[str, Any]:
    rules = [
        ("续航与能耗", "续航缩水/低温掉电", r"续航|掉电|电耗|里程|BMS|热泵|低温|冬天|缩水"),
        ("智能座舱", "车机系统卡顿", r"车机|中控屏|语音|系统|OTA|卡顿|黑屏|死机|导航|投屏"),
        ("智能驾驶", "智驾误报或退出", r"智驾|辅助驾驶|自动泊车|AEB|雷达|误报|摄像头|雨雾|制动"),
        ("舒适性与NVH", "NVH/悬架舒适性问题", r"座椅|腰疼|悬架|悬挂|风噪|胎噪|异响|NVH|震动|压缩机"),
        ("充电体验", "充电兼容性或效率不足", r"充电|快充|慢充|充电桩|兼容|中断|限流"),
        ("空间表现", "空间舒适表现", r"空间|后排|前排|乘坐|储物|后备箱|宽敞"),
    ]
    for aspect, opinion, pattern in rules:
        if re.search(pattern, text):
            sentiment = classify_sentiment(text)
            return {"aspect": aspect, "opinion": opinion, "sentiment": sentiment, "reason": f"评论中提到“{opinion}”相关体验。", "confidence": 0.82, "need_review": False}
    sentiment = classify_sentiment(text)
    return {"aspect": "其他", "opinion": text[:18] or "未提取观点", "sentiment": sentiment, "reason": "未命中明确质量方面，建议专家复核。", "confidence": 0.5, "need_review": True}


async def llm_absa(text: str, provider: str) -> dict[str, Any]:
    api_key = os.getenv("DEEPSEEK_API_KEY") or os.getenv("OPENAI_API_KEY")
    if not api_key or provider == "rule":
        return rule_absa(text)
    prompt = f"""你是新能源汽车感知质量方面级情感分析专家。只输出合法JSON。
评论文本：{text}
方面类别必须从以下列表选择：{",".join(ASPECTS)}
输出 JSON: {{"aspect":"","opinion":"","sentiment":"","reason":"","confidence":0.0,"need_review":false}}"""
    try:
        async with httpx.AsyncClient(timeout=30) as client:
            resp = await client.post(
                "https://api.deepseek.com/v1/chat/completions",
                headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
                json={"model": "deepseek-chat", "temperature": 0.2, "response_format": {"type": "json_object"}, "messages": [{"role": "user", "content": prompt}]},
            )
            resp.raise_for_status()
            content = resp.json()["choices"][0]["message"]["content"]
            payload = json.loads(re.search(r"\{.*\}", content, re.S).group(0))
            aspect = payload.get("aspect") if payload.get("aspect") in ASPECTS else "其他"
            confidence = max(0, min(1, float(payload.get("confidence", 0.5))))
            return {
                "aspect": aspect,
                "opinion": str(payload.get("opinion") or "未提取观点"),
                "sentiment": "positive" if "正" in str(payload.get("sentiment")) else "negative" if "负" in str(payload.get("sentiment")) else "neutral",
                "reason": str(payload.get("reason") or "模型未返回明确原因。"),
                "confidence": confidence,
                "need_review": bool(payload.get("need_review")) or confidence < 0.65 or aspect == "其他",
            }
    except Exception:
        return rule_absa(text)


@app.post("/api/comments/upload")
async def upload_comments(file: UploadFile = File(...), text_column: str | None = Form(default=None)):
    data = parse_upload(file.filename or "comments.csv", await file.read())
    detected, confident = choose_text_column(data)
    chosen = text_column or detected
    if not chosen or not confident and not text_column:
        return {"needs_column_selection": True, "columns": list(data[0].keys()) if data else [], "detected_text_column": detected, "preview_rows": data[:5]}
    with db() as conn:
        conn.execute("delete from comments")
        conn.execute("delete from absa_results")
        for idx, row in enumerate(data, start=1):
            raw_text = str(row.get(chosen, "")).strip()
            if not raw_text:
                continue
            conn.execute("insert or replace into comments values (?,?,?,?,?)", (f"c{idx}", raw_text, pick(row, ["来源", "source", "platform"], "用户上传"), pick(row, ["车型", "model", "vehicle_model"], "上传车型"), pick(row, ["时间", "日期", "date", "time"], "")))
    return {"needs_column_selection": False, "detected_text_column": chosen, "count": len(rows("select * from comments")), "preview": rows("select * from comments limit 100")}


@app.get("/api/comments/preview")
def comments_preview():
    return {"items": rows("select * from comments limit 100")}


@app.post("/api/absa/run")
async def run_absa(req: AbsaRunRequest):
    comments = rows("select * from comments")
    with db() as conn:
        conn.execute("delete from absa_results")
        for item in comments:
            result = rule_absa(item["raw_text"]) if req.mode == "rule" else await llm_absa(item["raw_text"], req.provider)
            conn.execute("insert or replace into absa_results values (?,?,?,?,?,?,?,?)", (item["comment_id"], result["aspect"], result["opinion"], result["sentiment"], result["reason"], result["confidence"], int(result["need_review"]), req.mode))
    return {"total": len(comments), "success": len(rows("select * from absa_results where need_review=0")), "need_review": len(rows("select * from absa_results where need_review=1"))}


@app.get("/api/absa/results")
def absa_results(need_review: bool | None = None, limit: int = 200, offset: int = 0):
    where = "" if need_review is None else "where a.need_review=?"
    params: tuple[Any, ...] = (() if need_review is None else (int(need_review),)) + (limit, offset)
    return {"items": rows(f"select c.*, a.* from comments c join absa_results a using(comment_id) {where} limit ? offset ?", params)}


ISSUE_PATTERNS = [
    ("冬季续航衰减", "续航与能耗", r"续航|掉电|电耗|里程|BMS|热泵|低温|冬天|缩水"),
    ("车机系统卡顿", "智能座舱", r"车机|中控|屏|语音|系统|OTA|卡顿|黑屏|死机|导航|投屏"),
    ("NVH异常", "舒适性与NVH", r"悬架|悬挂|风噪|胎噪|异响|NVH|震动|座椅|压缩机"),
    ("智驾误报与退出", "智能驾驶", r"智驾|辅助驾驶|自动泊车|AEB|雷达|误报|摄像头|雨|雾|制动"),
    ("充电兼容性与效率不足", "充电体验", r"充电|快充|慢充|充电桩|兼容|中断|限流"),
]


def issue_for(item: dict[str, Any]) -> tuple[str, str]:
    text = f"{item.get('aspect','')} {item.get('opinion','')} {item.get('reason','')} {item.get('raw_text','')}"
    for name, aspect, pattern in ISSUE_PATTERNS:
        if re.search(pattern, text):
            return name, aspect
    return f"{item.get('aspect') or '其他'}体验问题", item.get("aspect") or "其他"


def intensity(text: str, sentiment: str) -> float:
    if sentiment != "negative":
        return 0
    if re.search(r"垃圾|完全不能用|崩溃", text):
        return 1.0
    if re.search(r"严重|很差|受不了|频繁", text):
        return 0.8
    if re.search(r"明显|经常|缩水|卡顿|异响|误报", text):
        return 0.6
    return 0.4 if re.search(r"有点|偶尔|一两秒", text) else 0.6


def normalize(values: list[float]) -> list[float]:
    if not values:
        return []
    mn, mx = min(values), max(values)
    return [0.5 if mx == mn else (v - mn) / (mx - mn) for v in values]


@app.post("/api/diagnosis/generate")
def generate_diagnosis(req: DiagnosisRequest):
    data = rows("select c.*, a.* from comments c join absa_results a using(comment_id) where a.need_review=0")
    total = len(data) or 1
    grouped: dict[str, list[dict[str, Any]]] = {}
    for item in data:
        issue, aspect = issue_for(item)
        grouped.setdefault(f"{issue}::{aspect}", []).append(item)
    base = []
    for idx, (key, items) in enumerate(grouped.items(), start=1):
        issue, aspect = key.split("::")
        neg = sum(1 for i in items if i["sentiment"] == "negative")
        pos = sum(1 for i in items if i["sentiment"] == "positive")
        attention = len(items) / total * 100
        dissatisfaction = neg / len(items) * 100
        inten = sum(intensity(f"{i['opinion']} {i['reason']} {i['raw_text']}", i["sentiment"]) for i in items) / max(1, neg) * 100
        pos_ratio = pos / len(items) * 100
        kano = "Attractive" if pos_ratio >= 50 and dissatisfaction < 20 else "Must-be" if dissatisfaction >= 70 and pos_ratio <= 20 else "One-dimensional"
        if aspect in ["安全配置", "续航与能耗", "售后服务与交付", "操控与底盘"] and dissatisfaction >= 50:
            kano = "Must-be"
        base.append({"id": f"issue-{idx}", "issue": issue, "aspect": aspect, "items": items, "A": attention, "D": dissatisfaction, "I": inten, "kano": kano, "neg": neg, "pos": pos})
    zA, zD, zI = normalize([b["A"] for b in base]), normalize([b["D"] for b in base]), normalize([b["I"] for b in base])
    factors = {"Must-be": 1.2, "One-dimensional": 1.0, "Attractive": 0.8}
    with db() as conn:
        conn.execute("delete from issue_summary")
        for idx, item in enumerate(base):
            fda = (0.3 * zA[idx] + 0.5 * zD[idx] + 0.2 * zI[idx]) * 100
            factor = factors[item["kano"]]
            pi = fda * factor
            evidence = [i["raw_text"] for i in item["items"][:3]]
            conn.execute("insert or replace into issue_summary values (?,?,?,?,?,?,?,?,?,?,?,?,?,?)", (item["id"], item["issue"], item["aspect"], len(item["items"]), item["neg"], item["pos"], round(item["A"], 1), round(item["D"], 1), round(item["I"], 1), item["kano"], round(fda, 1), factor, round(pi, 1), json.dumps(evidence, ensure_ascii=False)))
    return diagnosis_results()


@app.get("/api/diagnosis/results")
def diagnosis_results():
    items = rows("select * from issue_summary order by final_PI desc")
    return {"items": items, "bubble": items, "kano": rows("select kano_type, count(*) count from issue_summary group by kano_type"), "top": items[:10]}


QFD_MAP = [
    ("冬季续航衰减", "电池低温放电性能", 9, ["低温", "冬天", "掉电", "续航缩水", "电池"], "动力电池系统"),
    ("冬季续航衰减", "电池热管理效率", 9, ["低温", "热泵", "电池加热", "空调"], "热管理系统"),
    ("冬季续航衰减", "BMS低温控制策略", 9, ["BMS", "预热", "标定"], "三电控制系统"),
    ("车机系统卡顿", "座舱芯片算力", 9, ["SoC", "芯片", "多任务"], "智能座舱"),
    ("车机系统卡顿", "系统内存管理", 9, ["内存", "卡顿", "黑屏"], "智能座舱软件"),
    ("车机系统卡顿", "软件响应时间", 9, ["响应", "卡顿", "延迟"], "智能座舱软件"),
    ("NVH异常", "悬架结构设计", 9, ["悬架", "悬挂", "震动"], "底盘与悬架系统"),
    ("NVH异常", "空气弹簧供应质量", 9, ["空气悬架", "压缩机", "异响"], "底盘与悬架系统"),
    ("智驾误报与退出", "传感器感知精度", 9, ["雷达", "摄像头", "感知"], "智能驾驶感知系统"),
    ("智驾误报与退出", "算法识别阈值", 9, ["误报", "退出", "制动"], "智能驾驶算法"),
    ("充电兼容性与效率不足", "充电协议兼容性", 9, ["充电桩", "兼容", "协议"], "充电系统"),
]


@app.post("/api/qfd/generate")
def generate_qfd(req: QfdRequest):
    issues = rows("select * from issue_summary order by final_PI desc limit ?", (req.top_n,))
    max_pi = max([i["final_PI"] for i in issues] or [1])
    confidence = {r["comment_id"]: r["confidence"] for r in rows("select * from absa_results")}
    avg_conf = sum(confidence.values()) / len(confidence) if confidence else 0.8
    with db() as conn:
        conn.execute("delete from qfd_results")
        for issue in issues:
            for mapped_issue, feature, base, keywords, module in QFD_MAP:
                if mapped_issue != issue["issue_name"] and issue["aspect"] not in mapped_issue:
                    continue
                evidence = " ".join(json.loads(issue["evidence_json"] or "[]"))
                hits = [kw for kw in keywords if kw in evidence or kw in issue["issue_name"]]
                keyword_match = 1.2 if len(hits) >= 2 else 1.0 if hits else 0.8
                pi_factor = issue["final_PI"] / max_pi
                score = base * keyword_match * avg_conf * pi_factor
                explanation = f"base_relation={base}, keyword_match={keyword_match}, confidence={avg_conf:.2f}, pi_factor={pi_factor:.2f}"
                conn.execute("insert into qfd_results (issue_id, engineering_feature, base_relation, keyword_match, confidence_factor, pi_factor, relation_score, module, explanation) values (?,?,?,?,?,?,?,?,?)", (issue["issue_id"], feature, base, keyword_match, avg_conf, pi_factor, round(score, 2), module, explanation))
    return qfd_results()


@app.get("/api/qfd/results")
def qfd_results():
    items = rows("select q.*, i.issue_name, i.aspect from qfd_results q left join issue_summary i using(issue_id) order by relation_score desc")
    importance = rows("select engineering_feature, module, round(sum(relation_score),2) importance from qfd_results group by engineering_feature, module order by importance desc")
    return {"items": items, "matrix": items, "feature_importance": importance}


ENTERPRISES = {
    "动力电池系统": [("宁德时代", "主协同主体", "低温放电性能优化、电芯材料参数复核、低温测试数据共享"), ("比亚迪弗迪电池", "协同主体", "低温测试与BMS策略协同")],
    "热管理系统": [("三花智控", "主协同主体", "热泵效率优化、冷却管路参数协同、热管理标定测试"), ("银轮股份", "协同主体", "换热效率和台架测试协同")],
    "三电控制系统": [("整车厂三电部门", "主协同主体", "BMS低温策略优化、OTA标定升级、联合台架测试"), ("华为数字能源", "协同主体", "能量管理策略协同")],
    "智能座舱": [("高通", "主协同主体", "算力适配、芯片负载测试、多任务性能优化"), ("芯驰科技", "协同主体", "座舱域控算力验证")],
    "智能座舱软件": [("华为车BU", "主协同主体", "系统响应优化、内存管理优化、OTA修复"), ("科大讯飞", "协同主体", "语音响应与离线语音优化")],
    "底盘与悬架系统": [("拓普集团", "主协同主体", "结构参数复核、NVH测试、装配工艺优化"), ("保隆科技", "协同主体", "空气弹簧供应质量复核")],
    "智能驾驶感知系统": [("禾赛科技", "主协同主体", "传感器标定、雨雾场景数据集补充、识别阈值优化"), ("速腾聚创", "协同主体", "点云质量与鲁棒性测试")],
    "智能驾驶算法": [("Momenta", "主协同主体", "识别阈值优化、场景数据训练"), ("地平线", "协同主体", "感知算法与芯片适配")],
    "充电系统": [("特来电", "主协同主体", "协议兼容测试、充电桩适配、异常中断数据共享"), ("星星充电", "协同主体", "充电桩适配与运营数据共享")],
}


@app.post("/api/supply-chain/generate")
def generate_supply_chain(req: SupplyChainRequest):
    features = rows("select engineering_feature, module, sum(relation_score) importance from qfd_results group by engineering_feature, module")
    role_weight = {"主协同主体": 1.0, "协同主体": 0.7, "支持主体": 0.4}
    with db() as conn:
        conn.execute("delete from supply_chain_results")
        for feature in features:
            for enterprise, role, method in ENTERPRISES.get(feature["module"], []):
                score = feature["importance"] * role_weight.get(role, 0.7)
                reason = f"基于QFD工程特征“{feature['engineering_feature']}”的重要度{feature['importance']:.2f}生成的候选协同主体推荐。"
                conn.execute("insert into supply_chain_results (engineering_feature, module, enterprise_name, role_type, collaboration_score, collaboration_method, recommendation_reason, model_suggested) values (?,?,?,?,?,?,?,0)", (feature["engineering_feature"], feature["module"], enterprise, role, round(score, 2), method, reason))
    return supply_chain_results()


@app.get("/api/supply-chain/results")
def supply_chain_results():
    return {"items": rows("select * from supply_chain_results order by collaboration_score desc")}


@app.get("/api/case/full-chain")
def case_full_chain():
    comments = rows("select * from comments limit 5")
    absa = rows("select c.raw_text, a.* from comments c join absa_results a using(comment_id) limit 5")
    issues = rows("select * from issue_summary order by final_PI desc limit 5")
    qfd = rows("select * from qfd_results order by relation_score desc limit 8")
    scm = rows("select * from supply_chain_results order by collaboration_score desc limit 8")
    return {"comments": comments, "absa": absa, "issues": issues, "qfd": qfd, "supply_chain": scm, "summary": "评论→ABSA→Kano-FDA→QFD→供应链协同→报告的完整链路已生成。"}


@app.post("/api/reports/generate")
def generate_report(req: ReportRequest):
    report_id = f"report-{int(time.time())}"
    path = REPORT_DIR / f"{report_id}.xlsx"
    try:
        from openpyxl import Workbook
        wb = Workbook()
        sheets = [
            ("数据概况", [{"comments": len(rows("select * from comments")), "absa": len(rows("select * from absa_results"))}]),
            ("ABSA分析结果", rows("select c.raw_text, a.* from comments c join absa_results a using(comment_id)")),
            ("Top质量问题", rows("select * from issue_summary order by final_PI desc limit ?", (req.top_n,))),
            ("QFD工程特征", rows("select * from qfd_results order by relation_score desc")),
            ("潜在协同企业", rows("select * from supply_chain_results order by collaboration_score desc")),
        ]
        for idx, (title, data) in enumerate(sheets):
            ws = wb.active if idx == 0 else wb.create_sheet(title)
            ws.title = title
            if data:
                headers = list(data[0].keys())
                ws.append(headers)
                for row in data:
                    ws.append([row.get(h) for h in headers])
        wb.save(path)
    except Exception:
        path = REPORT_DIR / f"{report_id}.json"
        path.write_text(json.dumps(case_full_chain(), ensure_ascii=False, indent=2), encoding="utf-8")
    with db() as conn:
        conn.execute("insert or replace into report_records values (?,?,?,?)", (report_id, req.report_type, time.strftime("%Y-%m-%d %H:%M:%S"), str(path)))
    return {"report_id": report_id, "download_url": f"/api/reports/download/{report_id}", "file_path": str(path)}


@app.get("/api/reports")
def list_reports():
    return {"items": rows("select * from report_records order by created_at desc")}


@app.get("/api/reports/download/{report_id}")
def download_report(report_id: str):
    record = rows("select * from report_records where report_id=?", (report_id,))
    if not record:
        raise HTTPException(404, "report not found")
    return FileResponse(record[0]["file_path"], filename=Path(record[0]["file_path"]).name)
